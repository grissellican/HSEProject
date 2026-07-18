from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as django_login, logout as django_logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password
from django.utils import timezone
from django.db.models import Avg, Count, Q
from .models import Course, Module, Material, Assignment, Submission, LiveSession, Question, Choice, QuestionResponse, ExamAttempt, ModuleAnnouncement, ModuleLink, ModuleForum, ForumReply, SubmissionFile, QuestionResponseFile, Cohort
from .forms import UserForm, CourseForm, ModuleForm, MaterialForm, AssignmentForm, GradeForm, LiveSessionForm, TeacherProfileForm, StudentSubmissionForm, CohortForm, CohortCloseForm

User = get_user_model()

# --- KPI DINÁMICOS Y ADAPTATIVOS ---
def get_dashboard_metrics(section='general'):
    """Calcula las métricas dinámicas basadas en la sección actual y estado del servidor"""
    total_users = User.objects.count()
    total_courses = Course.objects.count()
    active_teachers = User.objects.filter(role='teacher', is_active=True).count()
    
    # Verificación simulada de salud del sitio (Estado de Servidor)
    server_status = "Excelente (100%)" if total_users > 0 else "Inactivo"

    metrics = {
        'total_users': total_users,
        'active_teachers': active_teachers,
        'total_courses': total_courses,
        'server_status': server_status,
    }

    # KPIs contextuales adaptados según requerimiento
    if section == 'docentes':
        metrics.update({
            'kpi_1_label': 'Docentes Totales',
            'kpi_1_val': User.objects.filter(role='teacher').count(),
            'kpi_2_label': 'Docentes Activos',
            'kpi_2_val': active_teachers,
        })
    elif section == 'estudiantes':
        metrics.update({
            'kpi_1_label': 'Estudiantes Totales',
            'kpi_1_val': User.objects.filter(role='student').count(),
            'kpi_2_label': 'Estudiantes Activos',
            'kpi_2_val': User.objects.filter(role='student', is_active=True).count(),
        })
    elif section == 'usuarios': # Sección de Administradores
        metrics.update({
            'kpi_1_label': 'Administradores',
            'kpi_1_val': User.objects.filter(role='admin').count(),
            'kpi_2_label': 'Admins Activos',
            'kpi_2_val': User.objects.filter(role='admin', is_active=True).count(),
        })
    elif section == 'cursos':
        metrics.update({
            'kpi_1_label': 'Cursos Totales',
            'kpi_1_val': total_courses,
            'kpi_2_label': 'Aulas Activas',
            'kpi_2_val': Course.objects.filter(is_active=True).count(),
        })
    else: # Panel General
        metrics.update({
            'kpi_1_label': 'Usuarios Totales',
            'kpi_1_val': total_users,
            'kpi_2_label': 'Docentes Activos',
            'kpi_2_val': active_teachers,
        })
    return metrics


# --- LOGIN / LOGOUT CON REDIRECCIÓN DE ROLES ---
def login_view(request):
    if request.user.is_authenticated:
        return redirect_dashboard_by_role(request.user)

    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        remember = request.POST.get('remember') == 'on'

        if not email or not password:
            messages.error(request, "Por favor, complete todos los campos.")
            return render(request, 'auth/login.html')

        user = authenticate(request, username=email, password=password)

        if user is not None:
            if not user.is_active:
                messages.error(request, "Su cuenta se encuentra inactiva. Contacte al soporte de HSE.")
                return render(request, 'auth/login.html')
            
            django_login(request, user)
            
            if not remember:
                request.session.set_expiry(0)
            
            return redirect_dashboard_by_role(user)
        else:
            messages.error(request, "Las credenciales no coinciden con nuestros registros.")

    return render(request, 'auth/login.html')

def logout_view(request):
    django_logout(request)
    return redirect('login')

def redirect_dashboard_by_role(user):
    if user.role == 'admin':
        return redirect('admin_panel_general')
    elif user.role == 'teacher':
        return redirect('teacher_dashboard')
    else:
        return redirect('student_dashboard')


# --- VISTAS DEL ESTUDIANTE (Para redirección exitosa) ---
@login_required
def student_dashboard(request):
    if request.user.role != 'student':
        return redirect_dashboard_by_role(request.user)
    # El estudiante ve únicamente las clases en las que está matriculado
    my_courses = Course.objects.filter(students=request.user, is_active=True)
    return render(request, 'dashboards/student/student.html', {'courses': my_courses})


# ============================================================
# ===  DASHBOARD DEL DOCENTE — VISTAS COMPLETAS  ===
# ============================================================

def _teacher_required(view_func):
    """Decorador que verifica que el usuario sea docente o administrador."""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.role not in ['teacher', 'admin']:
            return redirect_dashboard_by_role(request.user)
        return view_func(request, *args, **kwargs)
    return wrapper


def _sidebar_context(request):
    """Retorna el contexto necesario para el sidebar compartido del docente."""
    if request.user.role == 'admin':
        return {'pending_sidebar': [], 'pending_sidebar_count': 0}
        
    pending_qs = Submission.objects.filter(
        assignment__module__course__teacher=request.user,
        score__isnull=True
    ).select_related('student', 'assignment').order_by('submitted_at')[:8]
    return {
        'pending_sidebar': pending_qs,
        'pending_sidebar_count': Submission.objects.filter(
            assignment__module__course__teacher=request.user,
            score__isnull=True
        ).count(),
    }



def _get_teacher_object(request, model, teacher_lookup, **kwargs):
    """Obtiene un objeto verificando que pertenezca al docente actual o si es admin en modo supervisión."""
    if request.user.role != 'admin':
        kwargs[teacher_lookup] = request.user
    return get_object_or_404(model, **kwargs)


def _get_teacher_course(request, course_id):
    """Obtiene un curso verificando que pertenezca al docente actual o si es admin en modo supervisión."""
    if request.user.role == 'admin':
        course = get_object_or_404(Course, id=course_id)
        return course
    return get_object_or_404(Course, id=course_id, teacher=request.user)


# --- PANEL PRINCIPAL DEL DOCENTE ---
@_teacher_required
def teacher_dashboard(request):
    my_courses = Course.objects.filter(teacher=request.user, is_active=True)
    
    # KPIs del docente
    total_students = 0
    for c in my_courses:
        total_students += c.students.count()
    
    pending_reviews = Submission.objects.filter(
        assignment__module__course__teacher=request.user,
        score__isnull=True
    ).count()
    
    upcoming_sessions = LiveSession.objects.filter(
        course__teacher=request.user,
        scheduled_date__gte=timezone.now().date()
    ).order_by('scheduled_date', 'start_time')[:5]
    
    context = {
        'section': 'dashboard',
        'sidebar_active': 'dashboard',
        'courses': my_courses,
        'total_courses': my_courses.count(),
        'total_students': total_students,
        'pending_reviews': pending_reviews,
        'upcoming_sessions': upcoming_sessions,
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/teacher/teacher.html', context)


# --- DETALLE DE UN CURSO ---
@_teacher_required
def teacher_course_detail(request, course_id):
    course = _get_teacher_course(request, course_id)
    modules = course.modules.filter(cohort__isnull=True).prefetch_related('materials', 'assignments')
    live_sessions = course.live_sessions.all().order_by('-scheduled_date', '-start_time')
    students = course.students.all()
    
    # Contar tareas y evaluaciones pendientes de revisión en este curso
    pending_in_course = Submission.objects.filter(
        assignment__module__course=course,
        score__isnull=True
    ).count()
    
    context = {
        'section': 'curso_detalle',
        'sidebar_active': 'curso',
        'course': course,
        'modules': modules,
        'live_sessions': live_sessions,
        'students': students,
        'pending_in_course': pending_in_course,
        'all_courses': Course.objects.filter(teacher=request.user, is_active=True),
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/teacher/teacher_course_detail.html', context)


@_teacher_required
def teacher_edit_syllabus(request, course_id):
    course = _get_teacher_course(request, course_id)
    if not course.allow_teacher_edit_syllabus:
        messages.error(request, "No tienes permiso para editar el sílabo y la evaluación de este curso.")
        return redirect('teacher_course_detail', course_id=course.id)
        
    if request.method == 'POST':
        form = CourseSyllabusTeacherForm(request.POST, request.FILES, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, "Sílabo y sistema de evaluación actualizados exitosamente.")
            return redirect('teacher_course_detail', course_id=course.id)
        else:
            messages.error(request, "Hubo un error al actualizar la información. Revisa los datos e inténtalo de nuevo.")
    else:
        form = CourseSyllabusTeacherForm(instance=course)
        
    context = {
        'section': 'curso_detalle',
        'sidebar_active': 'curso',
        'course': course,
        'form': form,
        'all_courses': Course.objects.filter(teacher=request.user, is_active=True),
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/teacher/teacher_edit_syllabus.html', context)


@login_required
def course_syllabus(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    # Verificar acceso (si es profe o estudiante del curso, o admin)
    
    if request.user.role == 'student':
        cohort = Cohort.objects.filter(course=course, students=request.user).first()
        if cohort and cohort.status in ['completed', 'archived']:
            modules = course.modules.filter(is_visible=True, cohort=cohort)
        else:
            modules = course.modules.filter(is_visible=True, cohort__isnull=True)
    else:
        modules = course.modules.filter(cohort__isnull=True)
        
    # Preparar contexto
    context = {
        'course': course,
        'modules': modules
    }
    
    if request.user.role == 'student':
        context.update(_student_sidebar_context(request))
        context['sidebar_active'] = 'curso'
        context['section'] = 'curso_detalle'
        template = 'dashboards/shared/course_syllabus.html'
    elif request.user.role == 'teacher':
        context.update(_sidebar_context(request))
        context['sidebar_active'] = 'curso'
        context['section'] = 'curso_detalle'
        template = 'dashboards/shared/course_syllabus.html'
    else:
        template = 'dashboards/shared/course_syllabus.html'
        
    return render(request, template, context)


@login_required
def course_evaluation(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    
    context = {
        'course': course,
    }
    
    if request.user.role == 'student':
        context.update(_student_sidebar_context(request))
        context['sidebar_active'] = 'curso'
        context['section'] = 'curso_detalle'
    elif request.user.role == 'teacher':
        context.update(_sidebar_context(request))
        context['sidebar_active'] = 'curso'
        context['section'] = 'curso_detalle'
        
    return render(request, 'dashboards/shared/course_evaluation.html', context)


# --- CRUD DE MÓDULOS ---
@_teacher_required
def teacher_module_create(request, course_id):
    course = _get_teacher_course(request, course_id)
    if request.method == 'POST':
        form = ModuleForm(request.POST)
        if form.is_valid():
            module = form.save(commit=False)
            module.course = course
            module.save()
            messages.success(request, f'Módulo "{module.title}" creado exitosamente.')
            return redirect('teacher_course_detail', course_id=course.id)
    else:
        # Auto-set order to next sequential number
        next_order = (course.modules.filter(cohort__isnull=True).count()) + 1
        form = ModuleForm(initial={'order': next_order})
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': 'Crear Nuevo Módulo',
        'course': course,
        'back_url': 'teacher_course_detail',
        'back_id': course.id,
    })


@_teacher_required
def teacher_module_update(request, module_id):
    module = _get_teacher_object(request, Module, 'course__teacher', id=module_id)
    if request.method == 'POST':
        form = ModuleForm(request.POST, instance=module)
        if form.is_valid():
            form.save()
            messages.success(request, f'Módulo "{module.title}" actualizado.')
            return redirect('teacher_course_detail', course_id=module.course.id)
    else:
        form = ModuleForm(instance=module)
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Editar Módulo: {module.title}',
        'course': module.course,
        'back_url': 'teacher_course_detail',
        'back_id': module.course.id,
    })


@_teacher_required
def teacher_module_delete(request, module_id):
    module = _get_teacher_object(request, Module, 'course__teacher', id=module_id)
    course_id = module.course.id
    module.delete()
    messages.success(request, 'Módulo eliminado correctamente.')
    return redirect('teacher_course_detail', course_id=course_id)

@_teacher_required
def teacher_module_toggle_visibility(request, module_id):
    module = _get_teacher_object(request, Module, 'course__teacher', id=module_id)
    module.is_visible = not module.is_visible
    module.save()
    status = "visible" if module.is_visible else "oculto"
    messages.success(request, f'Módulo "{module.title}" ahora está {status}.')
    return redirect('teacher_course_detail', course_id=module.course.id)


# --- CRUD DE MATERIALES ---
@_teacher_required
def teacher_material_create(request, module_id):
    module = _get_teacher_object(request, Module, 'course__teacher', id=module_id)
    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES)
        if form.is_valid():
            material = form.save(commit=False)
            material.module = module
            material.save()
            messages.success(request, f'Material "{material.title}" subido exitosamente.')
            return redirect('teacher_course_detail', course_id=module.course.id)
    else:
        form = MaterialForm()
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Subir Material al Módulo: {module.title}',
        'course': module.course,
        'back_url': 'teacher_course_detail',
        'back_id': module.course.id,
        'is_file_form': True,
    })


@_teacher_required
def teacher_material_detail(request, material_id):
    material = _get_teacher_object(request, Material, 'module__course__teacher', id=material_id
    )
    context = {
        'material': material,
        'course': material.module.course,
        'module': material.module,
        'sidebar_active': 'curso',
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/teacher/teacher_material_detail.html', context)


@_teacher_required
def teacher_material_delete(request, material_id):
    material = _get_teacher_object(request, Material, 'module__course__teacher', id=material_id)
    course_id = material.module.course.id
    material.delete()
    messages.success(request, 'Material eliminado correctamente.')
    return redirect('teacher_course_detail', course_id=course_id)

@_teacher_required
def teacher_material_toggle_visibility(request, material_id):
    material = _get_teacher_object(request, Material, 'module__course__teacher', id=material_id)
    material.is_visible = not material.is_visible
    material.save()
    status = "visible" if material.is_visible else "oculto"
    messages.success(request, f'Material "{material.title}" ahora está {status}.')
    return redirect('teacher_course_detail', course_id=material.module.course.id)


# --- CRUD DE TAREAS Y EVALUACIONES ---
@_teacher_required
def teacher_assignment_create(request, module_id):
    module = _get_teacher_object(request, Module, 'course__teacher', id=module_id)
    if request.method == 'POST':
        form = AssignmentForm(request.POST, request.FILES)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.module = module
            assignment.save()
            
            # Guardar config para cohorte activa
            from .models import CohortAssignmentConfig
            active_cohort = module.course.cohorts.filter(status='active').first()
            if active_cohort:
                CohortAssignmentConfig.objects.update_or_create(
                    cohort=active_cohort,
                    assignment=assignment,
                    defaults={
                        'due_date': assignment.due_date,
                        'is_visible': assignment.is_visible
                    }
                )
                
            messages.success(request, f'{assignment.get_assignment_type_display()} "{assignment.title}" creada exitosamente.')
            return redirect('teacher_course_detail', course_id=module.course.id)
    else:
        form = AssignmentForm()
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Nueva Tarea/Evaluación — {module.title}',
        'course': module.course,
        'back_url': 'teacher_course_detail',
        'back_id': module.course.id,
        'is_file_form': True,
    })


@_teacher_required
def teacher_assignment_update(request, assignment_id):
    assignment = _get_teacher_object(request, Assignment, 'module__course__teacher', id=assignment_id)
    if request.method == 'POST':
        form = AssignmentForm(request.POST, request.FILES, instance=assignment)
        if form.is_valid():
            assignment = form.save()
            
            # Actualizar config para cohorte activa
            from .models import CohortAssignmentConfig
            active_cohort = assignment.module.course.cohorts.filter(status='active').first()
            if active_cohort:
                CohortAssignmentConfig.objects.update_or_create(
                    cohort=active_cohort,
                    assignment=assignment,
                    defaults={
                        'due_date': assignment.due_date,
                        'is_visible': assignment.is_visible
                    }
                )
                
            messages.success(request, f'{assignment.get_assignment_type_display()} "{assignment.title}" actualizada.')
            return redirect('teacher_course_detail', course_id=assignment.module.course.id)
    else:
        form = AssignmentForm(instance=assignment)
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Editar: {assignment.title}',
        'course': assignment.module.course,
        'back_url': 'teacher_course_detail',
        'back_id': assignment.module.course.id,
        'is_file_form': True,
    })


@_teacher_required
def teacher_assignment_delete(request, assignment_id):
    assignment = _get_teacher_object(request, Assignment, 'module__course__teacher', id=assignment_id)
    course_id = assignment.module.course.id
    assignment.delete()
    messages.success(request, 'Tarea/Evaluación eliminada correctamente.')
    return redirect('teacher_course_detail', course_id=course_id)

@_teacher_required
def teacher_assignment_toggle_visibility(request, assignment_id):
    assignment = _get_teacher_object(request, Assignment, 'module__course__teacher', id=assignment_id)
    assignment.is_visible = not assignment.is_visible
    assignment.save()
    status = "visible" if assignment.is_visible else "oculta"
    messages.success(request, f'{assignment.get_assignment_type_display()} "{assignment.title}" ahora está {status}.')
    return redirect('teacher_course_detail', course_id=assignment.module.course.id)


# --- CRUD DE AVISOS ---
@_teacher_required
def teacher_announcement_create(request, module_id):
    from .forms import AnnouncementForm
    module = _get_teacher_object(request, Module, 'course__teacher', id=module_id)
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            ann = form.save(commit=False)
            ann.module = module
            # Asignar a cohorte activa automáticamente
            active_cohort = module.course.cohorts.filter(status='active').first()
            if active_cohort:
                ann.cohort = active_cohort
            ann.save()
            messages.success(request, f'Aviso "{ann.title}" creado exitosamente.')
            return redirect('teacher_course_detail', course_id=module.course.id)
    else:
        from django.utils import timezone
        form = AnnouncementForm(initial={'publish_date': timezone.now()})
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form, 'title': f'Nuevo Aviso — {module.title}', 'course': module.course,
        'back_url': 'teacher_course_detail', 'back_id': module.course.id,
    })

@_teacher_required
def teacher_announcement_update(request, pk):
    from .models import ModuleAnnouncement
    from .forms import AnnouncementForm
    ann = _get_teacher_object(request, ModuleAnnouncement, 'module__course__teacher', id=pk)
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, instance=ann)
        if form.is_valid():
            form.save()
            messages.success(request, f'Aviso "{ann.title}" actualizado.')
            return redirect('teacher_course_detail', course_id=ann.module.course.id)
    else:
        form = AnnouncementForm(instance=ann)
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form, 'title': f'Editar Aviso: {ann.title}', 'course': ann.module.course,
        'back_url': 'teacher_course_detail', 'back_id': ann.module.course.id,
    })

@_teacher_required
def teacher_announcement_delete(request, pk):
    from .models import ModuleAnnouncement
    ann = _get_teacher_object(request, ModuleAnnouncement, 'module__course__teacher', id=pk)
    course_id = ann.module.course.id
    ann.delete()
    messages.success(request, 'Aviso eliminado.')
    return redirect('teacher_course_detail', course_id=course_id)

# --- CRUD DE ENLACES ---
@_teacher_required
def teacher_link_create(request, module_id):
    from .forms import LinkForm
    module = _get_teacher_object(request, Module, 'course__teacher', id=module_id)
    if request.method == 'POST':
        form = LinkForm(request.POST)
        if form.is_valid():
            link = form.save(commit=False)
            link.module = module
            link.save()
            messages.success(request, f'Enlace "{link.title}" creado.')
            return redirect('teacher_course_detail', course_id=module.course.id)
    else:
        form = LinkForm()
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form, 'title': f'Nuevo Enlace — {module.title}', 'course': module.course,
        'back_url': 'teacher_course_detail', 'back_id': module.course.id,
    })

@_teacher_required
def teacher_link_update(request, pk):
    from .models import ModuleLink
    from .forms import LinkForm
    link = _get_teacher_object(request, ModuleLink, 'module__course__teacher', id=pk)
    if request.method == 'POST':
        form = LinkForm(request.POST, instance=link)
        if form.is_valid():
            form.save()
            messages.success(request, f'Enlace "{link.title}" actualizado.')
            return redirect('teacher_course_detail', course_id=link.module.course.id)
    else:
        form = LinkForm(instance=link)
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form, 'title': f'Editar Enlace: {link.title}', 'course': link.module.course,
        'back_url': 'teacher_course_detail', 'back_id': link.module.course.id,
    })

@_teacher_required
def teacher_link_delete(request, pk):
    from .models import ModuleLink
    link = _get_teacher_object(request, ModuleLink, 'module__course__teacher', id=pk)
    course_id = link.module.course.id
    link.delete()
    messages.success(request, 'Enlace eliminado.')
    return redirect('teacher_course_detail', course_id=course_id)

# --- CRUD DE FOROS ---
@_teacher_required
def teacher_forum_create(request, module_id):
    from .forms import ForumForm
    module = _get_teacher_object(request, Module, 'course__teacher', id=module_id)
    if request.method == 'POST':
        form = ForumForm(request.POST)
        if form.is_valid():
            forum = form.save(commit=False)
            forum.module = module
            # Asignar a cohorte activa automáticamente
            active_cohort = module.course.cohorts.filter(status='active').first()
            if active_cohort:
                forum.cohort = active_cohort
            forum.save()
            messages.success(request, f'Foro "{forum.title}" creado.')
            return redirect('teacher_course_detail', course_id=module.course.id)
    else:
        from django.utils import timezone
        form = ForumForm(initial={'start_date': timezone.now()})
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form, 'title': f'Nuevo Foro — {module.title}', 'course': module.course,
        'back_url': 'teacher_course_detail', 'back_id': module.course.id,
    })

@_teacher_required
def teacher_forum_update(request, pk):
    from .models import ModuleForum
    from .forms import ForumForm
    forum = _get_teacher_object(request, ModuleForum, 'module__course__teacher', id=pk)
    if request.method == 'POST':
        form = ForumForm(request.POST, instance=forum)
        if form.is_valid():
            form.save()
            messages.success(request, f'Foro "{forum.title}" actualizado.')
            return redirect('teacher_course_detail', course_id=forum.module.course.id)
    else:
        form = ForumForm(instance=forum)
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form, 'title': f'Editar Foro: {forum.title}', 'course': forum.module.course,
        'back_url': 'teacher_course_detail', 'back_id': forum.module.course.id,
    })

@_teacher_required
def teacher_forum_delete(request, pk):
    from .models import ModuleForum
    forum = _get_teacher_object(request, ModuleForum, 'module__course__teacher', id=pk)
    course_id = forum.module.course.id
    forum.delete()
    messages.success(request, 'Foro eliminado.')
    return redirect('teacher_course_detail', course_id=course_id)


@_teacher_required
def teacher_announcement_toggle_visibility(request, pk):
    from .models import ModuleAnnouncement
    ann = _get_teacher_object(request, ModuleAnnouncement, 'module__course__teacher', id=pk)
    ann.is_visible = not ann.is_visible
    ann.save()
    return redirect('teacher_course_detail', course_id=ann.module.course.id)

@_teacher_required
def teacher_link_toggle_visibility(request, pk):
    from .models import ModuleLink
    link = _get_teacher_object(request, ModuleLink, 'module__course__teacher', id=pk)
    link.is_visible = not link.is_visible
    link.save()
    return redirect('teacher_course_detail', course_id=link.module.course.id)

@_teacher_required
def teacher_forum_toggle_visibility(request, pk):
    from .models import ModuleForum
    forum = _get_teacher_object(request, ModuleForum, 'module__course__teacher', id=pk)
    forum.is_visible = not forum.is_visible
    forum.save()
    forum.save()
    return redirect('teacher_course_detail', course_id=forum.module.course.id)

@_teacher_required
def teacher_forum_detail(request, pk):
    from .forms import ForumReplyForm
    from .models import ModuleForum
    forum = _get_teacher_object(request, ModuleForum, 'module__course__teacher', id=pk)
    
    if request.method == 'POST':
        if not forum.is_active:
            messages.error(request, 'El foro está cerrado.')
            return redirect('teacher_forum_detail', pk=forum.id)
            
        form = ForumReplyForm(request.POST)
        if form.is_valid():
            reply = form.save(commit=False)
            reply.forum = forum
            reply.author = request.user
            
            parent_id = request.POST.get('parent_id')
            if parent_id:
                try:
                    from .models import ForumReply
                    parent_reply = ForumReply.objects.get(id=parent_id, forum=forum)
                    reply.parent = parent_reply
                except ForumReply.DoesNotExist:
                    pass
            
            reply.save()
            messages.success(request, 'Respuesta publicada exitosamente.')
            return redirect('teacher_forum_detail', pk=forum.id)
    else:
        form = ForumReplyForm()
        
    context = {
        'forum': forum,
        'replies': forum.replies.filter(parent__isnull=True).order_by('created_at').prefetch_related('nested_replies'),
        'form': form,
        'course': forum.module.course,
        'sidebar_active': 'dashboard',
    }
    return render(request, 'dashboards/teacher/teacher_forum_detail.html', context)

@_teacher_required
def teacher_forum_reply_delete(request, reply_id):
    from .models import ForumReply
    reply = _get_teacher_object(request, ForumReply, 'forum__module__course__teacher', id=reply_id)
    if request.method == 'POST':
        forum_id = reply.forum.id
        reply.delete()
        messages.success(request, 'Respuesta eliminada exitosamente.')
        return redirect('teacher_forum_detail', pk=forum_id)
    return redirect('teacher_dashboard')

# --- ENTREGAS Y CALIFICACIONES ---
@_teacher_required
def teacher_assignment_submissions(request, assignment_id):
    assignment = _get_teacher_object(request, Assignment, 'module__course__teacher', id=assignment_id)
    submissions = assignment.submissions.select_related('student').all()
    
    # Calcular estadísticas de calificación
    graded = submissions.filter(score__isnull=False)
    avg_score = graded.aggregate(avg=Avg('score'))['avg']
    avg_percentage = None
    if avg_score is not None and assignment.max_score > 0:
        avg_percentage = round((avg_score / assignment.max_score) * 100, 1)
    
    context = {
        'section': 'entregas',
        'sidebar_active': 'pending',
        'assignment': assignment,
        'course': assignment.module.course,
        'submissions': submissions,
        'total_submissions': submissions.count(),
        'graded_count': graded.count(),
        'pending_count': submissions.filter(score__isnull=True).count(),
        'avg_score': avg_score,
        'avg_percentage': avg_percentage,
        'all_courses': Course.objects.filter(teacher=request.user, is_active=True),
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/teacher/teacher_submissions.html', context)


@_teacher_required
def teacher_grade_submission(request, submission_id):
    submission = _get_teacher_object(request, Submission, 'assignment__module__course__teacher', id=submission_id)
    
    # Si es un examen, usamos la vista de calificación específica de exámenes
    if submission.assignment.assignment_type == 'examen_online':
        return teacher_grade_exam_submission(request, submission)
        
    if request.method == 'POST':
        form = GradeForm(request.POST, instance=submission, max_score=float(submission.assignment.max_score))
        if form.is_valid():
            sub = form.save(commit=False)
            sub.graded_at = timezone.now()
            sub.save()
            messages.success(request, f'Calificación registrada para {submission.student.first_name} {submission.student.lastname}.')
            return redirect('teacher_assignment_submissions', assignment_id=submission.assignment.id)
    else:
        form = GradeForm(instance=submission, max_score=float(submission.assignment.max_score))
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Calificar Entrega de {submission.student.first_name} {submission.student.lastname}',
        'course': submission.assignment.module.course,
        'submission': submission,
        'back_url': 'teacher_assignment_submissions',
        'back_id': submission.assignment.id,
    })


def teacher_grade_exam_submission(request, submission):
    """Vista dedicada a calificar envíos de exámenes."""
    responses = submission.question_responses.select_related('question', 'selected_choice').order_by('question__order')
    
    if request.method == 'POST':
        # Procesar calificaciones manuales para preguntas de texto o archivo
        total_score = 0
        for resp in responses:
            if resp.question.question_type in ['text', 'file']:
                score_str = request.POST.get(f'score_{resp.id}')
                feedback_str = request.POST.get(f'feedback_{resp.id}', '')
                if score_str:
                    try:
                        resp.score = float(score_str)
                        if resp.score > float(resp.question.points):
                            messages.warning(request, f'La puntuación ingresada para una pregunta excedía el máximo ({resp.question.points}) y fue ajustada.')
                            resp.score = float(resp.question.points)
                    except ValueError:
                        resp.score = 0
                resp.feedback = feedback_str
                resp.save()
            
            if resp.score is not None:
                total_score += float(resp.score)
        
        # Actualizar puntaje global de la submission
        submission.score = total_score
        submission.graded_at = timezone.now()
        submission.save()
        messages.success(request, f'Examen calificado exitosamente. Puntaje final: {total_score}')
        return redirect('teacher_assignment_submissions', assignment_id=submission.assignment.id)

    # Calcular puntajes auto-obtenidos y los máximos
    total_auto_score = sum(r.score for r in responses if r.question.question_type == 'multiple_choice' and r.score is not None)
    
    return render(request, 'dashboards/teacher/teacher_grade_exam.html', {
        'submission': submission,
        'responses': responses,
        'course': submission.assignment.module.course,
        'assignment': submission.assignment,
        'total_auto_score': total_auto_score,
    })


# --- CLASES EN VIVO ---
@_teacher_required
def teacher_live_session_create(request, course_id):
    course = _get_teacher_course(request, course_id)
    if request.method == 'POST':
        form = LiveSessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.course = course
            # Asignar a cohorte activa automáticamente
            active_cohort = course.cohorts.filter(status='active').first()
            if active_cohort:
                session.cohort = active_cohort
            session.save()
            messages.success(request, f'Sesión en vivo "{session.title}" programada exitosamente.')
            return redirect('teacher_course_detail', course_id=course.id)
    else:
        form = LiveSessionForm()
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': 'Programar Clase en Vivo',
        'course': course,
        'back_url': 'teacher_course_detail',
        'back_id': course.id,
    })


@_teacher_required
def teacher_live_session_delete(request, session_id):
    session = _get_teacher_object(request, LiveSession, 'course__teacher', id=session_id)
    course_id = session.course.id
    session.delete()
    messages.success(request, 'Sesión en vivo cancelada/eliminada.')
    return redirect('teacher_course_detail', course_id=course_id)

@_teacher_required
def teacher_live_session_update(request, session_id):
    session = _get_teacher_object(request, LiveSession, 'course__teacher', id=session_id)
    if request.method == 'POST':
        form = LiveSessionForm(request.POST, instance=session)
        if form.is_valid():
            form.save()
            messages.success(request, 'Clase en vivo actualizada exitosamente.')
            return redirect('teacher_course_detail', course_id=session.course.id)
    else:
        form = LiveSessionForm(instance=session)
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Editar Sesión: {session.title}',
        'course': session.course,
        'back_url': 'teacher_course_detail',
        'back_id': session.course.id,
    })


# --- ESTUDIANTES Y CALIFICACIONES DEL CURSO ---
@_teacher_required
def teacher_students_list(request, course_id):
    course = _get_teacher_course(request, course_id)
    cohort_id = request.GET.get('cohort_id')
    active_cohorts = Cohort.objects.filter(course=course, status='active')
    
    if cohort_id:
        current_cohort = get_object_or_404(Cohort, id=cohort_id, course=course, status='active')
    else:
        current_cohort = active_cohorts.first()
        
    if current_cohort:
        students = current_cohort.students.all()
    else:
        students = course.students.all()
    
    # Para cada estudiante, calcular su promedio de calificaciones en este curso
    student_grades = []
    all_assignments = Assignment.objects.filter(module__course=course)
    total_assignments = all_assignments.count()
    
    for student in students:
        if current_cohort:
            submissions = Submission.objects.filter(
                student=student,
                assignment__module__course=course,
                cohort=current_cohort
            )
        else:
            submissions = Submission.objects.filter(
                student=student,
                assignment__module__course=course
            )
        graded_submissions = submissions.filter(score__isnull=False)
        
        # Calcular promedio ponderado sobre el puntaje máximo
        total_score = 0
        total_max = 0
        for sub in graded_submissions:
            total_score += float(sub.score)
            total_max += float(sub.assignment.max_score)
        
        avg_percentage = round((total_score / total_max) * 100, 1) if total_max > 0 else None
        
        student_grades.append({
            'student': student,
            'submitted': submissions.count(),
            'graded': graded_submissions.count(),
            'total_assignments': total_assignments,
            'avg_percentage': avg_percentage,
            'total_score': round(total_score, 2),
            'total_max': round(total_max, 2),
        })
    
    context = {
        'section': 'estudiantes',
        'course': course,
        'student_grades': student_grades,
        'total_assignments': total_assignments,
        'all_courses': Course.objects.filter(teacher=request.user, is_active=True),
        'active_cohorts': active_cohorts,
        'current_cohort': current_cohort,
    }
    return render(request, 'dashboards/teacher/teacher_students.html', context)

@_teacher_required
def teacher_student_grades_detail(request, course_id, student_id):
    from django.utils import timezone
    course = _get_teacher_course(request, course_id)
    student = get_object_or_404(User, id=student_id)
    
    if student not in course.students.all():
        messages.error(request, 'El estudiante no pertenece a este curso.')
        return redirect('teacher_students_list', course_id=course.id)
    
    modules = course.modules.all().prefetch_related('assignments')
    
    module_data = []
    
    for module in modules:
        assignments_data = []
        for assignment in module.assignments.all():
            submission = Submission.objects.filter(student=student, assignment=assignment).first()
            
            status = 'No entregado'
            score_display = '—'
            
            if submission:
                if submission.score is not None:
                    status = 'Calificado'
                    score_display = f"{submission.score} / {assignment.max_score}"
                else:
                    status = 'Pendiente de calificar'
                    score_display = f"— / {assignment.max_score}"
            else:
                if assignment.due_date and timezone.now() > assignment.due_date:
                    status = 'Fuera de plazo'
                    score_display = f"0 / {assignment.max_score}"
                else:
                    status = 'Aún no calificado'
                    score_display = f"— / {assignment.max_score}"
                    
            assignments_data.append({
                'assignment': assignment,
                'submission': submission,
                'status': status,
                'score_display': score_display,
            })
            
        module_data.append({
            'module': module,
            'assignments_data': assignments_data
        })
        
    context = {
        'section': 'estudiantes',
        'course': course,
        'student': student,
        'module_data': module_data,
        'all_courses': Course.objects.filter(teacher=request.user, is_active=True),
    }
    return render(request, 'dashboards/teacher/teacher_student_grades.html', context)


# --- PERFIL DEL DOCENTE ---
@_teacher_required
def teacher_profile(request):
    user_id = request.GET.get('user_id')
    if request.user.role == 'admin' and user_id:
        user = get_object_or_404(User, id=user_id, role='teacher')
        is_admin_supervising = True
    else:
        user = request.user
        is_admin_supervising = False
        
    if request.method == 'POST' and not is_admin_supervising:
        form = TeacherProfileForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu perfil se ha actualizado correctamente.')
            return redirect('teacher_profile')
    else:
        form = TeacherProfileForm(instance=user)
        if is_admin_supervising:
            for field in form.fields.values():
                field.disabled = True
        
    context = {
        'section': 'perfil',
        'form': form,
        'profile_user': user,
        'is_admin_supervising': is_admin_supervising,
        'all_courses': Course.objects.filter(teacher=user, is_active=True),
    }
    return render(request, 'dashboards/teacher/teacher_profile.html', context)


# --- SECCIONES DEL PANEL ADMINISTRADOR ---

@login_required
def admin_platform_settings(request):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    
    from .models import PlatformSetting
    from .forms import PlatformSettingForm
    
    settings = PlatformSetting.get_settings()
    
    if request.method == 'POST':
        form = PlatformSettingForm(request.POST, request.FILES, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuración actualizada exitosamente.")
            return redirect('admin_platform_settings')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
            messages.error(request, "No se pudo actualizar la configuración. Revisa los errores.")
    else:
        form = PlatformSettingForm(instance=settings)
        
    context = {
        'section': 'configuracion',
        'form': form,
    }
    return render(request, 'dashboards/admin/admin_platform_settings.html', context)

@login_required
def admin_panel_general(request):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    context = get_dashboard_metrics('general')
    context.update({
        'section': 'general',
        'table_title': 'Panel Centralizado de Registros Recientes',
        'users_list': User.objects.all().order_by('-id')[:5],
    })
    return render(request, 'dashboards/admin.html', context)

@login_required
def admin_section_users(request):
    """Gestión de Administradores por parte del Superadmin"""
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    context = get_dashboard_metrics('usuarios')
    context.update({
        'section': 'usuarios',
        'table_title': 'Gestión de Administradores',
        'users_list': User.objects.filter(role='admin').order_by('-id'),
    })
    return render(request, 'dashboards/admin.html', context)

@login_required
def admin_section_teachers(request):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    
    query = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    users = User.objects.filter(role='teacher')
    
    if query:
        from django.db.models import Q
        users = users.filter(Q(first_name__icontains=query) | Q(lastname__icontains=query) | Q(email__icontains=query))
    if date_from:
        users = users.filter(date_joined__gte=date_from)
    if date_to:
        users = users.filter(date_joined__lte=date_to + ' 23:59:59')
        
    context = get_dashboard_metrics('docentes')
    context.update({
        'section': 'docentes',
        'table_title': 'Gestión de Personal Docente',
        'users_list': users.order_by('-id'),
        'search_query': query,
        'date_from': date_from,
        'date_to': date_to,
    })
    return render(request, 'dashboards/admin.html', context)

@login_required
def admin_section_students(request):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    
    query = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    users = User.objects.filter(role='student')
    
    if query:
        from django.db.models import Q
        users = users.filter(Q(first_name__icontains=query) | Q(lastname__icontains=query) | Q(email__icontains=query))
    if date_from:
        users = users.filter(date_joined__gte=date_from)
    if date_to:
        users = users.filter(date_joined__lte=date_to + ' 23:59:59')
        
    context = get_dashboard_metrics('estudiantes')
    context.update({
        'section': 'estudiantes',
        'table_title': 'Control de Alumnos Matriculados',
        'users_list': users.order_by('-id'),
        'search_query': query,
        'date_from': date_from,
        'date_to': date_to,
    })
    return render(request, 'dashboards/admin.html', context)

@login_required
def admin_section_courses(request):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    context = get_dashboard_metrics('cursos')
    context.update({
        'section': 'cursos',
        'table_title': 'Catálogo de Aulas Virtuales',
        'courses_list': Course.objects.all().order_by('-id'),
    })
    return render(request, 'dashboards/admin.html', context)
    
@login_required
def admin_course_detail(request, course_id):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    from .models import Course, Cohort
    course = get_object_or_404(Course, id=course_id)
    
    active_cohorts = Cohort.objects.filter(course=course, status='active').count()
    completed_cohorts = Cohort.objects.filter(course=course, status='completed').count()
    
    context = {
        'course': course,
        'active_cohorts_count': active_cohorts,
        'completed_cohorts_count': completed_cohorts,
        'sidebar_active': 'cursos',
        'section': 'cursos',
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/admin/admin_course_detail.html', context)


# --- OPERACIONES CRUD USUARIOS ---

@login_required
def user_create(request, target_role=None):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    initial_data = {}
    if target_role:
        initial_data['role'] = target_role
        
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = user.email
            if form.cleaned_data['password']:
                user.password = make_password(form.cleaned_data['password'])
            user.save()
            if 'enrolled_courses' in form.cleaned_data:
                user.enrolled_courses.set(form.cleaned_data['enrolled_courses'])
            messages.success(request, f'Cuenta registrada de manera exitosa.')
            
            if target_role == 'teacher': return redirect('admin_docentes')
            if target_role == 'student': return redirect('admin_estudiantes')
            return redirect('admin_usuarios')
    else:
        form = UserForm(initial=initial_data)
    
    return render(request, 'dashboards/form_modal.html', {
        'form': form, 
        'title': f'Registrar Nuevo {target_role.upper() if target_role else "Usuario"}'
    })

@login_required
def user_update(request, pk):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            u = form.save(commit=False)
            if form.cleaned_data['password']:
                u.password = make_password(form.cleaned_data['password'])
            u.save()
            if 'enrolled_courses' in form.cleaned_data:
                u.enrolled_courses.set(form.cleaned_data['enrolled_courses'])
            messages.success(request, 'Cuenta modificada de manera exitosa.')
            
            if u.role == 'teacher': return redirect('admin_docentes')
            if u.role == 'student': return redirect('admin_estudiantes')
            return redirect('admin_usuarios')
    else:
        form = UserForm(instance=user)
    return render(request, 'dashboards/form_modal.html', {'form': form, 'title': 'Modificar Cuenta de Usuario'})

@login_required
def user_delete(request, pk):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    user = get_object_or_404(User, pk=pk)
    role = user.role
    user.delete()
    messages.success(request, 'El registro ha sido removido del sistema.')
    
    if role == 'teacher': return redirect('admin_docentes')
    if role == 'student': return redirect('admin_estudiantes')
    return redirect('admin_usuarios')


# --- OPERACIONES CRUD CURSOS (Con guardado Many-To-Many de alumnos) ---

@login_required
def course_create(request):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES)
        if form.is_valid():
            course = form.save()
            messages.success(request, f'Aula virtual "{course.title}" dada de alta correctamente.')
            return redirect('admin_cursos')
    else:
        form = CourseForm()
    return render(request, 'dashboards/form_modal.html', {'form': form, 'title': 'Añadir Nueva Aula Virtual', 'is_file_form': True})

@login_required
def course_update(request, pk):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cambios en el aula "{course.title}" aplicados con éxito.')
            return redirect('admin_course_detail', course_id=course.id)
    else:
        form = CourseForm(instance=course)
    return render(request, 'dashboards/form_modal.html', {'form': form, 'title': f'Editar Aula: {course.title}', 'is_file_form': True})

@login_required
def course_students_update(request, pk):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    course = get_object_or_404(Course, pk=pk)
    from .forms import CourseStudentsForm
    if request.method == 'POST':
        form = CourseStudentsForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, f'Estudiantes actualizados para el aula "{course.title}".')
            return redirect('admin_course_detail', course_id=course.id)
    else:
        form = CourseStudentsForm(instance=course)
    return render(request, 'dashboards/form_modal.html', {'form': form, 'title': f'Matricular Alumnos: {course.title}'})

@login_required
def course_delete(request, pk):
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    course = get_object_or_404(Course, pk=pk)
    title = course.title
    course.delete()
    messages.success(request, f'El aula "{title}" ha sido eliminada por completo.')
    return redirect('admin_cursos')


# ============================================================
# ===  GESTIÓN DE EXÁMENES ONLINE  ===
# ============================================================

from .models import Question, Choice
from .forms import QuestionForm, ChoiceForm

@_teacher_required
def teacher_exam_questions(request, assignment_id):
    assignment = _get_teacher_object(request, Assignment, 'module__course__teacher', id=assignment_id, assignment_type='examen_online')
    questions = assignment.questions.prefetch_related('choices').all()
    
    total_points = sum(q.points for q in questions)
    
    ctx = {
        'sidebar_active': 'pending',
        'assignment': assignment,
        'course': assignment.module.course,
        'questions': questions,
        'total_points': total_points,
    }
    ctx.update(_sidebar_context(request))
    return render(request, 'dashboards/teacher/teacher_exam_questions.html', ctx)


@_teacher_required
def teacher_question_create(request, assignment_id):
    assignment = _get_teacher_object(request, Assignment, 'module__course__teacher', id=assignment_id, assignment_type='examen_online')
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.assignment = assignment
            question.save()
            messages.success(request, f'Pregunta creada exitosamente.')
            return redirect('teacher_exam_questions', assignment_id=assignment.id)
    else:
        next_order = assignment.questions.count() + 1
        form = QuestionForm(initial={'order': next_order})
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Nueva Pregunta para: {assignment.title}',
        'course': assignment.module.course,
        'back_url': 'teacher_exam_questions',
        'back_id': assignment.id,
    })


@_teacher_required
def teacher_question_update(request, question_id):
    question = _get_teacher_object(request, Question, 'assignment__module__course__teacher', id=question_id)
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, f'Pregunta actualizada.')
            return redirect('teacher_exam_questions', assignment_id=question.assignment.id)
    else:
        form = QuestionForm(instance=question)
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Editar Pregunta',
        'course': question.assignment.module.course,
        'back_url': 'teacher_exam_questions',
        'back_id': question.assignment.id,
    })


@_teacher_required
def teacher_question_delete(request, question_id):
    question = _get_teacher_object(request, Question, 'assignment__module__course__teacher', id=question_id)
    assignment_id = question.assignment.id
    question.delete()
    messages.success(request, 'Pregunta eliminada.')
    return redirect('teacher_exam_questions', assignment_id=assignment_id)


@_teacher_required
def teacher_choice_create(request, question_id):
    question = _get_teacher_object(request, Question, 'assignment__module__course__teacher', id=question_id, question_type='multiple_choice')
    if request.method == 'POST':
        form = ChoiceForm(request.POST)
        if form.is_valid():
            choice = form.save(commit=False)
            choice.question = question
            # Si es correcta, desmarcamos las demás correctas (asumiendo que solo hay 1 correcta en opciones simples, aunque esto es opcional)
            # if choice.is_correct:
            #     question.choices.update(is_correct=False)
            choice.save()
            messages.success(request, f'Opción añadida exitosamente.')
            return redirect('teacher_exam_questions', assignment_id=question.assignment.id)
    else:
        form = ChoiceForm()
        
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Nueva Opción para la pregunta: {question.text[:30]}...',
        'course': question.assignment.module.course,
        'back_url': 'teacher_exam_questions',
        'back_id': question.assignment.id,
    })


@_teacher_required
def teacher_choice_delete(request, choice_id):
    choice = _get_teacher_object(request, Choice, 'question__assignment__module__course__teacher', id=choice_id)
    assignment_id = choice.question.assignment.id
    choice.delete()
    messages.success(request, 'Opción eliminada.')
    return redirect('teacher_exam_questions', assignment_id=assignment_id)


# ============================================================
# ===  DASHBOARD DEL ESTUDIANTE — VISTAS COMPLETAS  ===
# ============================================================

def _student_required(view_func):
    """Decorador que verifica que el usuario sea estudiante."""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.role != 'student':
            return redirect_dashboard_by_role(request.user)
        return view_func(request, *args, **kwargs)
    return wrapper


def _student_sidebar_context(request):
    """Contexto compartido para el sidebar del estudiante."""
    # Filtrado a nivel Python para usar las configs de cohorte
    pending_count = 0
    assignments = Assignment.objects.filter(
        module__course__students=request.user,
        module__is_visible=True,
        is_visible=True
    ).exclude(
        submissions__student=request.user
    )
    for asg in assignments:
        if not asg.get_is_visible_for_user(request.user): continue
        due_date = asg.get_due_date_for_user(request.user)
        if due_date is None or due_date >= timezone.now():
            pending_count += 1
            
    return {'pending_count': pending_count}


def _get_student_course(request, course_id):
    """Obtiene un curso verificando que el estudiante esté matriculado o tenga acceso histórico por cohorte."""
    # Primero intentar acceso activo
    course = Course.objects.filter(id=course_id, students=request.user, is_active=True).first()
    if course:
        return course
    cohort = Cohort.objects.filter(
        course_id=course_id,
        students=request.user,
        status='completed'
    ).first()
    if cohort and not cohort.is_expired(request.user):
        return cohort.course
    # Si no tiene acceso, 404
    return get_object_or_404(Course, id=course_id, students=request.user, is_active=True)


# --- PANEL PRINCIPAL DEL ESTUDIANTE ---
@_student_required
def student_dashboard(request):
    my_courses = Course.objects.filter(students=request.user, is_active=True)
    
    completed_cohorts = Cohort.objects.filter(
        students=request.user,
        status='completed'
    ).select_related('course')
    
    # ¿Tiene el estudiante algún curso activo?
    has_active_courses = Cohort.objects.filter(students=request.user, status='active').exists()
    
    completed_courses = []
    for cohort in completed_cohorts:
        if not cohort.is_expired(request.user):
            completed_courses.append({
                'course': cohort.course,
                'cohort': cohort,
                'days_remaining': cohort.get_days_until_expiration(request.user),
                'is_paused': has_active_courses
            })
    
    context = {
        'courses': my_courses,
        'completed_courses': completed_courses,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student.html', context)


# --- CALIFICACIONES ---
@_student_required
def student_grades(request):
    active_cohorts = Cohort.objects.filter(students=request.user, status='active').values_list('course_id', flat=True)
    active_courses_ids = set(active_cohorts)
    
    completed_cohorts = Cohort.objects.filter(students=request.user, status='completed')
    completed_courses_ids = set()
    for c in completed_cohorts:
        if not c.is_expired(request.user):
            completed_courses_ids.add(c.course_id)
            
    valid_courses_ids = active_courses_ids.union(completed_courses_ids)
    
    submissions = Submission.objects.filter(
        student=request.user,
        score__isnull=False,
        assignment__module__course_id__in=valid_courses_ids
    ).select_related('assignment__module__course').order_by('assignment__module__course__title', '-graded_at')
    
    active_grades = {}
    completed_grades = {}
    
    for sub in submissions:
        course = sub.assignment.module.course
        if course.id in active_courses_ids:
            if course.id not in active_grades:
                active_grades[course.id] = {'course': course, 'submissions': []}
            active_grades[course.id]['submissions'].append(sub)
        elif course.id in completed_courses_ids:
            if course.id not in completed_grades:
                completed_grades[course.id] = {'course': course, 'submissions': []}
            completed_grades[course.id]['submissions'].append(sub)
    
    context = {
        'active_grades': active_grades.values(),
        'completed_grades': completed_grades.values(),
        'sidebar_active': 'grades',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_grades.html', context)


# --- ACTIVIDADES PENDIENTES ---
@_student_required
def student_pending(request):
    active_cohorts = Cohort.objects.filter(students=request.user, status='active').values_list('id', flat=True)
    
    pending_assignments = Assignment.objects.filter(
        module__course__students=request.user,
        module__is_visible=True,
        is_visible=True,
    ).exclude(
        submissions__student=request.user
    ).select_related('module__course')
    
    pending_forums = ModuleForum.objects.filter(
        module__course__students=request.user,
        module__is_visible=True,
        is_visible=True,
    ).filter(
        Q(cohort__in=active_cohorts) | Q(cohort__isnull=True)
    ).filter(
        Q(end_date__gte=timezone.now()) | Q(end_date__isnull=True)
    ).exclude(
        replies__author=request.user
    ).select_related('module__course')
    
    pending_items = []
    for asg in pending_assignments:
        if not asg.get_is_visible_for_user(request.user): continue
        due_date = asg.get_due_date_for_user(request.user)
        if due_date is None or due_date >= timezone.now():
            pending_items.append({'type': 'assignment', 'obj': asg, 'date': due_date})
            
    for forum in pending_forums:
        pending_items.append({'type': 'forum', 'obj': forum, 'date': forum.end_date})
        
    import datetime
    def sort_key(item):
        return item['date'] if item['date'] else (timezone.now() + datetime.timedelta(days=3650))
    pending_items.sort(key=sort_key)
    
    context = {
        'pending_items': pending_items,
        'sidebar_active': 'pending',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_pending.html', context)


# --- MIS ARCHIVOS ---
@_student_required
def student_files(request):
    submissions_with_files = SubmissionFile.objects.filter(
        submission__student=request.user
    ).select_related('submission__assignment__module__course').order_by('-uploaded_at')
    
    # Also include exam file responses
    exam_files = QuestionResponseFile.objects.filter(
        response__submission__student=request.user
    ).select_related(
        'response__submission__assignment__module__course', 'response__question'
    ).order_by('-uploaded_at')
    
    context = {
        'submissions_with_files': submissions_with_files,
        'exam_files': exam_files,
        'sidebar_active': 'files',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_files.html', context)


# --- BIENVENIDA DEL CURSO ---
@_student_required
def student_course_home(request, course_id):
    course = _get_student_course(request, course_id)
    context = {
        'course': course,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_course_home.html', context)


# --- DOCENTE DEL CURSO ---
@_student_required
def student_course_teacher(request, course_id):
    course = _get_student_course(request, course_id)
    context = {
        'course': course,
        'teacher': course.teacher,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_course_teacher.html', context)


# --- CLASES EN VIVO ---
@_student_required
def student_course_live(request, course_id):
    course = _get_student_course(request, course_id)
    student_cohort = Cohort.objects.filter(course=course, students=request.user).first()
    
    sessions = course.live_sessions.all()
    if student_cohort:
        sessions = sessions.filter(Q(cohort=student_cohort) | Q(cohort__isnull=True))
    else:
        sessions = sessions.filter(cohort__isnull=True)
        
    sessions = sessions.order_by('-scheduled_date', '-start_time')
    
    context = {
        'course': course,
        'sessions': sessions,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_course_live.html', context)


# --- MÓDULOS DEL CURSO ---
@_student_required
def student_course_modules(request, course_id):
    from django.db.models import Prefetch
    course = _get_student_course(request, course_id)
    student_cohort = Cohort.objects.filter(course=course, students=request.user).first()
    
    if student_cohort:
        if student_cohort.status in ['completed', 'archived']:
            announcements_qs = ModuleAnnouncement.objects.filter(cohort=student_cohort)
            forums_qs = ModuleForum.objects.filter(cohort=student_cohort)
            modules_qs = course.modules.filter(cohort=student_cohort)
        else:
            announcements_qs = ModuleAnnouncement.objects.filter(Q(cohort=student_cohort) | Q(cohort__isnull=True))
            forums_qs = ModuleForum.objects.filter(Q(cohort=student_cohort) | Q(cohort__isnull=True))
            modules_qs = course.modules.filter(cohort__isnull=True)
    else:
        announcements_qs = ModuleAnnouncement.objects.filter(cohort__isnull=True)
        forums_qs = ModuleForum.objects.filter(cohort__isnull=True)
        modules_qs = course.modules.filter(cohort__isnull=True)
        
    modules = modules_qs.filter(is_visible=True).prefetch_related(
        'materials', 'assignments', 'links',
        Prefetch('announcements', queryset=announcements_qs),
        Prefetch('forums', queryset=forums_qs)
    )
    
    context = {
        'course': course,
        'modules': modules,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_course_modules.html', context)


# --- DETALLE DE MATERIAL ---
@_student_required
def student_material_detail(request, material_id):
    material = get_object_or_404(
        Material,
        id=material_id,
        is_visible=True,
        module__is_visible=True,
        module__course__in=Course.objects.filter(
            Q(students=request.user) | Q(cohorts__students=request.user)
        )
    )
    context = {
        'material': material,
        'course': material.module.course,
        'module': material.module,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_material_detail.html', context)


# --- DETALLE DE AVISO ---
@_student_required
def student_announcement_detail(request, announcement_id):
    announcement = get_object_or_404(
        ModuleAnnouncement,
        id=announcement_id,
        is_visible=True,
        module__is_visible=True,
        module__course__in=Course.objects.filter(
            Q(students=request.user) | Q(cohorts__students=request.user)
        )
    )
    context = {
        'announcement': announcement,
        'course': announcement.module.course,
        'module': announcement.module,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_announcement_detail.html', context)


# --- DETALLE DE FORO ---
@_student_required
def student_forum_detail(request, forum_id):
    from .forms import ForumReplyForm
    forum = get_object_or_404(
        ModuleForum,
        id=forum_id,
        is_visible=True,
        module__is_visible=True,
        module__course__in=Course.objects.filter(
            Q(students=request.user) | Q(cohorts__students=request.user)
        )
    )
    
    if request.method == 'POST':
        if not forum.is_active:
            messages.error(request, 'El foro está cerrado.')
            return redirect('student_forum_detail', forum_id=forum.id)
            
        form = ForumReplyForm(request.POST)
        if form.is_valid():
            reply = form.save(commit=False)
            reply.forum = forum
            reply.author = request.user
            
            parent_id = request.POST.get('parent_id')
            if parent_id:
                try:
                    from .models import ForumReply
                    parent_reply = ForumReply.objects.get(id=parent_id, forum=forum)
                    reply.parent = parent_reply
                except ForumReply.DoesNotExist:
                    pass
                    
            reply.save()
            messages.success(request, 'Respuesta publicada exitosamente.')
            return redirect('student_forum_detail', forum_id=forum.id)
    else:
        form = ForumReplyForm()
        
    can_view_replies = True
    if forum.forum_type == 'cerrado' and forum.is_active:
        has_participated = forum.replies.filter(author=request.user).exists()
        if not has_participated:
            can_view_replies = False
            
    context = {
        'forum': forum,
        'replies': forum.replies.filter(parent__isnull=True).order_by('created_at').prefetch_related('nested_replies'),
        'can_view_replies': can_view_replies,
        'form': form,
        'course': forum.module.course,
        'module': forum.module,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_forum_detail.html', context)

@_student_required
def student_forum_reply_edit(request, reply_id):
    from .forms import ForumReplyForm
    from .models import ForumReply
    reply = get_object_or_404(
        ForumReply,
        id=reply_id,
        author=request.user,
        forum__is_visible=True,
        forum__module__is_visible=True
    )
    
    if not reply.forum.is_active:
        messages.error(request, 'El foro está cerrado. No puedes editar tu respuesta.')
        return redirect('student_forum_detail', forum_id=reply.forum.id)
        
    if request.method == 'POST':
        form = ForumReplyForm(request.POST, instance=reply)
        if form.is_valid():
            form.save()
            messages.success(request, 'Respuesta actualizada exitosamente.')
            return redirect('student_forum_detail', forum_id=reply.forum.id)
    else:
        form = ForumReplyForm(instance=reply)
        
    context = {
        'reply': reply,
        'forum': reply.forum,
        'form': form,
        'course': reply.forum.module.course,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_forum_reply_edit.html', context)



# --- DETALLE DE TAREA ---
@_student_required
def student_assignment_detail(request, assignment_id):
    assignment = get_object_or_404(
        Assignment,
        id=assignment_id,
        assignment_type='tarea',
        is_visible=True,
        module__is_visible=True,
        module__course__in=Course.objects.filter(
            Q(students=request.user) | Q(cohorts__students=request.user)
        )
    )
    submission = Submission.objects.filter(assignment=assignment, student=request.user).first()
    
    # Check if past due date
    can_submit = True
    is_past_due = False
    due_date = assignment.get_due_date_for_user(request.user)
    if due_date and timezone.now() > due_date:
        is_past_due = True
        can_submit = False
    if submission:
        if assignment.max_attempts > 0 and submission.attempts >= assignment.max_attempts:
            can_submit = False
    
    form = StudentSubmissionForm()
    
    context = {
        'assignment': assignment,
        'submission': submission,
        'can_submit': can_submit,
        'is_past_due': is_past_due,
        'due_date': due_date,
        'form': form,
        'course': assignment.module.course,
        'module': assignment.module,
        'sidebar_active': 'dashboard',
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_assignment_detail.html', context)


# --- ENTREGAR TAREA ---
@_student_required
def student_submit_assignment(request, assignment_id):
    assignment = get_object_or_404(
        Assignment,
        id=assignment_id,
        is_visible=True,
        module__is_visible=True,
        module__course__in=Course.objects.filter(
            Q(students=request.user) | Q(cohorts__students=request.user)
        )
    )
    
    sub = Submission.objects.filter(assignment=assignment, student=request.user).first()
    
    # No permitir si alcanzó el límite de intentos
    if sub and assignment.max_attempts > 0 and sub.attempts >= assignment.max_attempts:
        messages.warning(request, 'Has alcanzado el límite de intentos para esta actividad.')
        return redirect('student_assignment_detail', assignment_id=assignment_id)
    
    # No permitir si pasó la fecha
    due_date = assignment.get_due_date_for_user(request.user)
    if due_date and timezone.now() > due_date:
        messages.error(request, 'La fecha límite de entrega ha pasado.')
        return redirect('student_assignment_detail', assignment_id=assignment_id)
    
    if request.method == 'POST':
        form = StudentSubmissionForm(request.POST, request.FILES, instance=sub)
        if form.is_valid():
            submission_instance = form.save(commit=False)
            if not sub:
                submission_instance.assignment = assignment
                submission_instance.student = request.user
                submission_instance.attempts = 1
                
                # Assign to active cohort if exists
                active_cohort = Cohort.objects.filter(course=assignment.module.course, students=request.user, status='active').first()
                if active_cohort:
                    submission_instance.cohort = active_cohort
            else:
                submission_instance.attempts = sub.attempts + 1
                submission_instance.score = None
                submission_instance.graded_at = None
                # Borrar archivos anteriores
                submission_instance.files.all().delete()
            
            submission_instance.save()
            
            for f in request.FILES.getlist('file'):
                SubmissionFile.objects.create(submission=submission_instance, file=f)
                
            messages.success(request, '¡Entrega realizada con éxito!')
            
            # Redirect based on assignment type
            if assignment.assignment_type == 'tarea':
                return redirect('student_assignment_detail', assignment_id=assignment_id)
            else:
                return redirect('student_evaluation_detail', assignment_id=assignment_id)
        else:
            print("Form errors:", form.errors)
    
    messages.error(request, 'Error al procesar la entrega.')
    return redirect('student_assignment_detail', assignment_id=assignment_id)


# --- DETALLE DE EVALUACIÓN SIMPLE ---
@_student_required
def student_evaluation_detail(request, assignment_id):
    assignment = get_object_or_404(
        Assignment,
        id=assignment_id,
        assignment_type='evaluacion',
        is_visible=True,
        module__is_visible=True,
        module__course__in=Course.objects.filter(
            Q(students=request.user) | Q(cohorts__students=request.user)
        )
    )
    submission = Submission.objects.filter(assignment=assignment, student=request.user).first()
    
    # Check disclaimer acceptance via session
    accepted_key = f'eval_accepted_{assignment_id}'
    has_accepted = request.session.get(accepted_key, False)
    
    can_submit = True
    is_past_due = False
    due_date = assignment.get_due_date_for_user(request.user)
    
    student_cohort = Cohort.objects.filter(course=assignment.module.course, students=request.user).first()
    is_frozen_cohort = student_cohort and student_cohort.status in ['completed', 'archived']
    
    if is_frozen_cohort:
        can_submit = False
    else:
        if due_date and timezone.now() > due_date:
            is_past_due = True
            can_submit = False
        if submission:
            if assignment.max_attempts > 0 and submission.attempts >= assignment.max_attempts:
                can_submit = False
    
    form = StudentSubmissionForm()
    
    context = {
        'assignment': assignment,
        'submission': submission,
        'has_accepted': has_accepted,
        'can_submit': can_submit,
        'is_past_due': is_past_due,
        'due_date': due_date,
        'form': form,
        'course': assignment.module.course,
        'module': assignment.module,
        'sidebar_active': 'dashboard',
        'is_frozen_cohort': is_frozen_cohort,
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_evaluation_detail.html', context)


# --- ACEPTAR DISCLAIMER DE EVALUACIÓN ---
@_student_required
def student_evaluation_accept(request, assignment_id):
    if request.method == 'POST':
        request.session[f'eval_accepted_{assignment_id}'] = True
    return redirect('student_evaluation_detail', assignment_id=assignment_id)


# --- DETALLE DE EXAMEN ONLINE ---
@_student_required
def student_exam_detail(request, assignment_id):
    assignment = get_object_or_404(
        Assignment,
        id=assignment_id,
        assignment_type='examen_online',
        is_visible=True,
        module__is_visible=True,
        module__course__in=Course.objects.filter(
            Q(students=request.user) | Q(cohorts__students=request.user)
        )
    )
    submission = Submission.objects.filter(assignment=assignment, student=request.user).first()
    attempt = None
    if submission:
        attempt = ExamAttempt.objects.filter(submission=submission).first()
        
    is_past_due = False
    due_date = assignment.get_due_date_for_user(request.user)
    if due_date and timezone.now() > due_date:
        is_past_due = True
        
    can_retake = False
    remaining_attempts = None
    
    student_cohort = Cohort.objects.filter(course=assignment.module.course, students=request.user).first()
    is_frozen_cohort = student_cohort and student_cohort.status in ['completed', 'archived']
    
    if not is_frozen_cohort:
        if attempt and attempt.is_completed and not is_past_due:
            if assignment.max_attempts == 0:
                can_retake = True
            elif submission.attempts < assignment.max_attempts:
                can_retake = True
                remaining_attempts = assignment.max_attempts - submission.attempts
    
    questions = assignment.questions.all()
    
    context = {
        'assignment': assignment,
        'submission': submission,
        'attempt': attempt,
        'is_past_due': is_past_due,
        'due_date': due_date,
        'can_retake': can_retake,
        'remaining_attempts': remaining_attempts,
        'total_questions': questions.count(),
        'course': assignment.module.course,
        'module': assignment.module,
        'sidebar_active': 'dashboard',
        'is_frozen_cohort': is_frozen_cohort,
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_exam_detail.html', context)


# --- EMPEZAR EXAMEN ONLINE ---
@_student_required
def student_exam_start(request, assignment_id):
    assignment = get_object_or_404(
        Assignment,
        id=assignment_id,
        assignment_type='examen_online',
        module__is_visible=True,
        module__course__in=Course.objects.filter(
            Q(students=request.user) | Q(cohorts__students=request.user)
        )
    )
    
    # No permitir si ya tiene un intento incompleto, o si agotó sus intentos
    existing = Submission.objects.filter(assignment=assignment, student=request.user).first()
    if existing:
        attempt = ExamAttempt.objects.filter(submission=existing).first()
        if attempt and not attempt.is_completed:
            if assignment.show_all_questions:
                return redirect('student_exam_all', assignment_id=assignment_id)
            # Redirigir a la pregunta actual
            return redirect('student_exam_question', assignment_id=assignment_id, q=attempt.current_question_index)
        
        # If they have an existing attempt, but it's completed
        if attempt and attempt.is_completed:
            if assignment.max_attempts > 0 and existing.attempts >= assignment.max_attempts:
                messages.info(request, 'Ya has agotado tus intentos para este examen.')
                return redirect('student_exam_detail', assignment_id=assignment_id)
    
    if request.method == 'POST':
        import traceback
        try:
            # Check due date
            due_date = assignment.get_due_date_for_user(request.user)
            if due_date and timezone.now() > due_date:
                messages.error(request, 'La fecha límite ha pasado.')
                return redirect('student_exam_detail', assignment_id=assignment_id)
            
            if existing:
                # Reusing the existing submission for a new attempt
                existing.attempts += 1
                existing.score = None
                existing.feedback = ''
                existing.graded_at = None
                existing.save()
                
                # Delete old attempt and responses
                if 'attempt' in locals() and attempt:
                    attempt.delete()
                
                existing.question_responses.all().delete()
                submission = existing
            else:
                # Create a new submission
                submission = Submission.objects.create(
                    assignment=assignment,
                    student=request.user
                )
                
                active_cohort = Cohort.objects.filter(course=assignment.module.course, students=request.user, status='active').first()
                if active_cohort:
                    submission.cohort = active_cohort
                    submission.save()
            
            import random
            # Mezclar preguntas
            q_ids = list(assignment.questions.values_list('id', flat=True))
            random.shuffle(q_ids)
            
            # Mezclar alternativas de preguntas de opción múltiple
            c_orders = {}
            for question in assignment.questions.filter(question_type='multiple_choice'):
                c_ids = list(question.choices.values_list('id', flat=True))
                random.shuffle(c_ids)
                c_orders[str(question.id)] = c_ids
                
            ExamAttempt.objects.create(
                submission=submission,
                current_question_index=0,
                question_order=q_ids,
                choice_orders=c_orders
            )
            if assignment.show_all_questions:
                return redirect('student_exam_all', assignment_id=assignment_id)
            return redirect('student_exam_question', assignment_id=assignment_id, q=0)
        except Exception as e:
            print("ERROR IN EXAM START:", flush=True)
            traceback.print_exc()
            raise
    
    return redirect('student_exam_detail', assignment_id=assignment_id)

# --- TODAS LAS PREGUNTAS DEL EXAMEN (Misma página) ---
@_student_required
def student_exam_all(request, assignment_id):
    assignment = get_object_or_404(
        Assignment,
        id=assignment_id,
        assignment_type='examen_online',
        module__course__students=request.user
    )
    if not assignment.show_all_questions:
        return redirect('student_exam_detail', assignment_id=assignment_id)
        
    submission = get_object_or_404(Submission, assignment=assignment, student=request.user)
    attempt = get_object_or_404(ExamAttempt, submission=submission)
    
    if attempt.is_completed:
        return redirect('student_exam_detail', assignment_id=assignment_id)
        
    from django.utils import timezone
    import datetime
    
    q_order = attempt.question_order
    
    if request.method == 'POST':
        # Check if due date is passed with a 1-minute grace period
        if assignment.due_date and timezone.now() > assignment.due_date + datetime.timedelta(minutes=1):
            messages.error(request, 'El tiempo del examen ha expirado.')
            return redirect('student_exam_finish', assignment_id=assignment_id)
            
        for q_id in q_order:
            question = get_object_or_404(Question, id=q_id, assignment=assignment)
            response, created = QuestionResponse.objects.get_or_create(
                submission=submission,
                question=question
            )
            
            if question.question_type == 'multiple_choice':
                choice_id = request.POST.get(f'choice_{q_id}')
                if choice_id:
                    try:
                        choice = Choice.objects.get(id=choice_id, question=question)
                        response.selected_choice = choice
                        if choice.is_correct:
                            response.score = question.points
                        else:
                            response.score = 0
                    except Choice.DoesNotExist:
                        pass
            elif question.question_type == 'text':
                response.text_answer = request.POST.get(f'text_{q_id}', '')
            elif question.question_type == 'file':
                response.save()
                files = request.FILES.getlist(f'file_{q_id}')
                for f in files:
                    QuestionResponseFile.objects.create(response=response, file=f)
                    
            response.save()
            
        return redirect('student_exam_finish', assignment_id=assignment_id)
        
    # GET: Prepare all questions
    prepared_questions = []
    for idx, q_id in enumerate(q_order):
        question = get_object_or_404(Question, id=q_id, assignment=assignment)
        existing_response = QuestionResponse.objects.filter(submission=submission, question=question).first()
        
        item = {
            'question': question,
            'question_num': idx + 1,
            'existing_response': existing_response,
        }
        
        if question.question_type == 'text':
            from .forms import ExamTextAnswerForm
            item['text_form'] = ExamTextAnswerForm(
                initial={'text_answer': existing_response.text_answer if existing_response else ''},
                prefix=f'text_{q_id}'
            )
        elif question.question_type == 'multiple_choice':
            choice_ids = attempt.choice_orders.get(str(question.id), [])
            choices_qs = question.choices.filter(id__in=choice_ids)
            choices_dict = {c.id: c for c in choices_qs}
            item['choices'] = [choices_dict[cid] for cid in choice_ids if cid in choices_dict]
            
        prepared_questions.append(item)
        
    context = {
        'assignment': assignment,
        'prepared_questions': prepared_questions,
        'course': assignment.module.course,
        'sidebar_active': 'dashboard',
        'due_date_iso': assignment.due_date.isoformat() if assignment.due_date else None,
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_exam_all.html', context)



# --- PREGUNTA DEL EXAMEN ---
@_student_required
def student_exam_question(request, assignment_id, q):
    assignment = get_object_or_404(
        Assignment,
        id=assignment_id,
        assignment_type='examen_online',
        module__course__students=request.user
    )
    submission = get_object_or_404(Submission, assignment=assignment, student=request.user)
    attempt = get_object_or_404(ExamAttempt, submission=submission)
    
    if attempt.is_completed:
        return redirect('student_exam_detail', assignment_id=assignment_id)
    
    q_order = attempt.question_order
    total = len(q_order)
    
    if q < 0 or q >= total:
        return redirect('student_exam_finish', assignment_id=assignment_id)
        
    # Enforce navigation rules
    if q > attempt.current_question_index:
        return redirect('student_exam_question', assignment_id=assignment_id, q=attempt.current_question_index)
        
    if not assignment.allow_backtracking and q < attempt.current_question_index:
        return redirect('student_exam_question', assignment_id=assignment_id, q=attempt.current_question_index)
    
    question_id = q_order[q]
    question = get_object_or_404(Question, id=question_id, assignment=assignment)
    existing_response = QuestionResponse.objects.filter(submission=submission, question=question).first()
    
    if request.method == 'POST':
        from django.utils import timezone
        import datetime
        
        # Check if due date is passed with a 1-minute grace period
        if assignment.due_date and timezone.now() > assignment.due_date + datetime.timedelta(minutes=1):
            messages.error(request, 'El tiempo del examen ha expirado.')
            return redirect('student_exam_finish', assignment_id=assignment_id)
            
        # Save response
        response, created = QuestionResponse.objects.get_or_create(
            submission=submission,
            question=question
        )
        
        if question.question_type == 'multiple_choice':
            choice_id = request.POST.get('choice')
            if choice_id:
                try:
                    choice = Choice.objects.get(id=choice_id, question=question)
                    response.selected_choice = choice
                    # Auto-grade multiple choice
                    if choice.is_correct:
                        response.score = question.points
                    else:
                        response.score = 0
                except Choice.DoesNotExist:
                    pass
        elif question.question_type == 'text':
            response.text_answer = request.POST.get('text_answer', '')
        elif question.question_type == 'file':
            response.save() # save first to get id
            files = request.FILES.getlist('file_answer')
            for f in files:
                QuestionResponseFile.objects.create(response=response, file=f)
        
        response.save()
        
        # Advance to next question only if we are at the furthest question
        if q == attempt.current_question_index:
            attempt.current_question_index = q + 1
            attempt.save()
        
        next_q = q + 1
        
        if next_q >= total:
            return redirect('student_exam_finish', assignment_id=assignment_id)
        return redirect('student_exam_question', assignment_id=assignment_id, q=next_q)
    
    text_form = None
    choices_list = []
    
    if question.question_type == 'text':
        from .forms import ExamTextAnswerForm
        text_form = ExamTextAnswerForm(initial={'text_answer': existing_response.text_answer if existing_response else ''})
    elif question.question_type == 'multiple_choice':
        choice_ids = attempt.choice_orders.get(str(question.id), [])
        choices_qs = question.choices.filter(id__in=choice_ids)
        choices_dict = {c.id: c for c in choices_qs}
        choices_list = [choices_dict[cid] for cid in choice_ids if cid in choices_dict]
        
    context = {
        'assignment': assignment,
        'question': question,
        'choices': choices_list,
        'question_num': q + 1,
        'total_questions': total,
        'existing_response': existing_response,
        'text_form': text_form,
        'course': assignment.module.course,
        'sidebar_active': 'dashboard',
        'due_date_iso': assignment.due_date.isoformat() if assignment.due_date else None,
    }
    context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/student/student_exam_question.html', context)


# --- FINALIZAR EXAMEN ---
@_student_required
def student_exam_finish(request, assignment_id):
    assignment = get_object_or_404(
        Assignment,
        id=assignment_id,
        assignment_type='examen_online',
        module__course__in=Course.objects.filter(
            Q(students=request.user) | Q(cohorts__students=request.user)
        )
    )
    submission = get_object_or_404(Submission, assignment=assignment, student=request.user)
    attempt = get_object_or_404(ExamAttempt, submission=submission)
    
    if not attempt.is_completed:
        attempt.is_completed = True
        attempt.finished_at = timezone.now()
        attempt.save()
        
        # Auto-calculate total score from graded responses (multiple choice)
        responses = submission.question_responses.all()
        auto_score = sum(r.score for r in responses if r.score is not None)
        # Only set score if all questions are auto-gradeable
        all_auto = all(r.question.question_type == 'multiple_choice' for r in responses)
        if all_auto and responses.exists():
            submission.score = auto_score
            submission.graded_at = timezone.now()
            submission.save()
        
        messages.success(request, '¡Examen finalizado exitosamente!')
    
    return redirect('student_exam_detail', assignment_id=assignment_id)


# --- SÍLABO Y EVALUACIÓN (VISTAS COMPARTIDAS Y DOCENTE) ---

def course_syllabus(request, course_id):
    if request.user.role == 'student':
        course = get_object_or_404(Course, id=course_id, students=request.user)
    elif request.user.role == 'teacher':
        course = _get_teacher_course(request, course_id)
    else:
        return redirect('login')
        
    context = {
        'course': course,
        'sidebar_active': 'dashboard',
    }
    if request.user.role == 'student':
        context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/shared/course_syllabus.html', context)

def course_evaluation(request, course_id):
    if request.user.role == 'student':
        course = get_object_or_404(Course, id=course_id, students=request.user)
    elif request.user.role == 'teacher':
        course = _get_teacher_course(request, course_id)
    else:
        return redirect('login')
        
    context = {
        'course': course,
        'evaluation_images': course.evaluation_images.all(),
        'sidebar_active': 'dashboard',
    }
    if request.user.role == 'student':
        context.update(_student_sidebar_context(request))
    return render(request, 'dashboards/shared/course_evaluation.html', context)

@_teacher_required
def teacher_edit_syllabus(request, course_id):
    course = _get_teacher_course(request, course_id)
    if not course.allow_teacher_edit_syllabus:
        messages.error(request, 'No tienes permisos para editar el sílabo de este curso.')
        return redirect('teacher_course_detail', course_id=course.id)
        
    from .forms import SyllabusTeacherForm, SyllabusUnitFormSet
    
    if request.method == 'POST':
        form = SyllabusTeacherForm(request.POST, instance=course)
        unit_formset = SyllabusUnitFormSet(request.POST, instance=course)
        
        if form.is_valid() and unit_formset.is_valid():
            form.save()
            unit_formset.save()
            messages.success(request, 'Sílabo actualizado correctamente.')
            return redirect('teacher_course_detail', course_id=course.id)
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = SyllabusTeacherForm(instance=course)
        unit_formset = SyllabusUnitFormSet(instance=course)
        
    context = {
        'course': course,
        'form': form,
        'unit_formset': unit_formset,
        'sidebar_active': 'dashboard',
    }
    return render(request, 'dashboards/teacher/teacher_edit_syllabus.html', context)

@_teacher_required
def teacher_edit_evaluation(request, course_id):
    course = _get_teacher_course(request, course_id)
    if not course.allow_teacher_edit_syllabus:
        messages.error(request, 'No tienes permisos para editar el sistema de evaluación.')
        return redirect('teacher_course_detail', course_id=course.id)
        
    from .forms import EvaluationSystemTeacherForm, EvaluationImageForm
    
    if request.method == 'POST' and 'update_description' in request.POST:
        form = EvaluationSystemTeacherForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, 'Descripción de la evaluación actualizada.')
            return redirect('teacher_edit_evaluation', course_id=course.id)
    else:
        form = EvaluationSystemTeacherForm(instance=course)
        
    if request.method == 'POST' and 'upload_image' in request.POST:
        img_form = EvaluationImageForm(request.POST, request.FILES)
        if img_form.is_valid():
            img = img_form.save(commit=False)
            img.course = course
            img.save()
            messages.success(request, 'Imagen añadida correctamente.')
            return redirect('teacher_edit_evaluation', course_id=course.id)
    else:
        img_form = EvaluationImageForm()
        
    context = {
        'course': course,
        'form': form,
        'img_form': img_form,
        'images': course.evaluation_images.all(),
        'sidebar_active': 'dashboard',
    }
    return render(request, 'dashboards/teacher/teacher_edit_evaluation.html', context)

@_teacher_required
def teacher_evaluation_image_delete(request, image_id):
    from .models import EvaluationImage
    img = _get_teacher_object(request, EvaluationImage, 'course__teacher', id=image_id)
    course_id = img.course.id
    if not img.course.allow_teacher_edit_syllabus:
        messages.error(request, 'No tienes permisos.')
        return redirect('teacher_course_detail', course_id=course_id)
        
    img.delete()
    messages.success(request, 'Imagen eliminada.')
    return redirect('teacher_edit_evaluation', course_id=course_id)

def _admin_required(view_func):
    """Decorador que verifica que el usuario sea administrador."""
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.role != 'admin':
            return redirect_dashboard_by_role(request.user)
        return view_func(request, *args, **kwargs)
    return wrapper

@_admin_required
def admin_edit_syllabus(request, course_id):
    course = get_object_or_404(Course, id=course_id)
        
    from .forms import SyllabusTeacherForm, SyllabusUnitFormSet
    
    if request.method == 'POST':
        form = SyllabusTeacherForm(request.POST, instance=course)
        unit_formset = SyllabusUnitFormSet(request.POST, instance=course)
        if form.is_valid() and unit_formset.is_valid():
            form.save()
            unit_formset.save()
            messages.success(request, 'Sílabo actualizado correctamente.')
            return redirect('admin_cursos')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = SyllabusTeacherForm(instance=course)
        unit_formset = SyllabusUnitFormSet(instance=course)
        
    context = {
        'course': course,
        'form': form,
        'unit_formset': unit_formset,
        'sidebar_active': 'cursos',
    }
    return render(request, 'dashboards/admin/admin_edit_syllabus.html', context)

@_admin_required
def admin_edit_evaluation(request, course_id):
    course = get_object_or_404(Course, id=course_id)
        
    from .forms import EvaluationSystemTeacherForm, EvaluationImageForm
    
    if request.method == 'POST' and 'update_description' in request.POST:
        form = EvaluationSystemTeacherForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, 'Descripción de la evaluación actualizada.')
            return redirect('admin_edit_evaluation', course_id=course.id)
    else:
        form = EvaluationSystemTeacherForm(instance=course)
        
    if request.method == 'POST' and 'upload_image' in request.POST:
        img_form = EvaluationImageForm(request.POST, request.FILES)
        if img_form.is_valid():
            img = img_form.save(commit=False)
            img.course = course
            img.save()
            messages.success(request, 'Imagen añadida correctamente.')
            return redirect('admin_edit_evaluation', course_id=course.id)
    else:
        img_form = EvaluationImageForm()
        
    context = {
        'course': course,
        'form': form,
        'img_form': img_form,
        'images': course.evaluation_images.all(),
        'sidebar_active': 'cursos',
    }
    return render(request, 'dashboards/admin/admin_edit_evaluation.html', context)

@_admin_required
def admin_evaluation_image_delete(request, image_id):
    from .models import EvaluationImage
    img = get_object_or_404(EvaluationImage, id=image_id)
    course_id = img.course.id
        
    img.delete()
    messages.success(request, 'Imagen eliminada.')
    return redirect('admin_edit_evaluation', course_id=course_id)


def custom_404(request, exception):
    return render(request, '404.html', status=404)

# ==========================================
# ASISTENCIAS (DOCENTE Y ADMINISTRADOR)
# ==========================================

@_teacher_required
def teacher_download_attendance_template(request, course_id):
    import openpyxl
    from django.http import HttpResponse
    
    course = _get_teacher_course(request, course_id)
    students = course.students.all().order_by('first_name', 'last_name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"
    
    # Headers
    ws.append(["ID", "Estudiante", "Estado (P/A/T)", "Observaciones"])
    
    # Data
    for student in students:
        ws.append([student.id, student.get_full_name(), "P", ""])
        
    # Formatting
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Plantilla_Asistencia_{course.title}.xlsx"'
    wb.save(response)
    return response

@_teacher_required
def teacher_attendance_upload(request, course_id):
    from .forms import AttendanceUploadForm
    from .models import AttendanceRegister, AttendanceRecord
    import openpyxl
    
    course = _get_teacher_course(request, course_id)
    
    if request.method == 'POST':
        form = AttendanceUploadForm(request.POST, request.FILES, course=course)
        if form.is_valid():
            file = form.cleaned_data['file']
            date = form.cleaned_data['date']
            description = form.cleaned_data['description']
            cohort = form.cleaned_data['cohort']
            
            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                
                if ws.max_row <= 1:
                    messages.error(request, 'El archivo Excel está vacío o no tiene el formato correcto.')
                    return redirect('teacher_attendance_upload', course_id=course.id)
                
                register = AttendanceRegister.objects.create(
                    course=course,
                    cohort=cohort,
                    date=date,
                    description=description,
                    uploaded_by=request.user
                )
                
                cohort_students = cohort.students.all()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue
                    
                    try:
                        student_id = int(row[0])
                        status_str = str(row[2]).strip().upper() if len(row) > 2 and row[2] else 'P'
                        observations = str(row[3]) if len(row) > 3 and row[3] else ''
                        
                        status_map = {'P': 'present', 'A': 'absent', 'T': 'late'}
                        status = status_map.get(status_str, 'present')
                        
                        student = cohort_students.filter(id=student_id).first()
                        if student:
                            AttendanceRecord.objects.create(
                                register=register,
                                student=student,
                                status=status,
                                observations=observations
                            )
                    except ValueError:
                        continue
                
                messages.success(request, f'Asistencia registrada exitosamente para la fecha {date}.')
                return redirect('teacher_attendance_upload', course_id=course.id)
                
            except Exception as e:
                messages.error(request, f'Error al procesar el archivo Excel: {str(e)}')
                return redirect('teacher_attendance_upload', course_id=course.id)
    else:
        form = AttendanceUploadForm(course=course)
        
    registers = course.attendance_registers.all().prefetch_related('records')
        
    context = {
        'section': 'curso_detalle',
        'course': course,
        'form': form,
        'registers': registers,
        'sidebar_active': 'curso',
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/teacher/teacher_attendance_upload.html', context)

@_admin_required
def admin_attendance_reports(request):
    from .models import Course, User
    courses = Course.objects.all()
    students = User.objects.filter(role='student').order_by('first_name', 'last_name')
    
    context = {
        'courses': courses,
        'students': students,
        'sidebar_active': 'asistencias',
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/admin/admin_attendance_reports.html', context)

@_admin_required
def admin_attendance_course_detail(request, course_id):
    from .models import Course, AttendanceRecord, Cohort
    course = get_object_or_404(Course, id=course_id)
    
    cohort_id = request.GET.get('cohort_id')
    all_cohorts = Cohort.objects.filter(course=course).order_by('-id')
    
    if cohort_id:
        current_cohort = get_object_or_404(Cohort, id=cohort_id, course=course)
    else:
        current_cohort = all_cohorts.filter(status='active').first() or all_cohorts.first()
        
    if current_cohort:
        registers = course.attendance_registers.filter(cohort=current_cohort)
        students = current_cohort.students.all()
    else:
        registers = course.attendance_registers.all()
        students = course.students.all()
    
    students_data = []
    total_classes = registers.count()
    
    for student in students:
        if current_cohort:
            records = AttendanceRecord.objects.filter(student=student, register__course=course, register__cohort=current_cohort)
        else:
            records = AttendanceRecord.objects.filter(student=student, register__course=course)
        presents = records.filter(status='present').count()
        absents = records.filter(status='absent').count()
        lates = records.filter(status='late').count()
        
        attended = presents + lates
        percentage = (attended / total_classes * 100) if total_classes > 0 else 0
        
        students_data.append({
            'student': student,
            'presents': presents,
            'absents': absents,
            'lates': lates,
            'percentage': round(percentage, 1),
        })
        
    students_data.sort(key=lambda x: x['student'].get_full_name())
    
    context = {
        'course': course,
        'registers': registers,
        'students_data': students_data,
        'all_cohorts': all_cohorts,
        'current_cohort': current_cohort,
        'total_classes': total_classes,
        'sidebar_active': 'asistencias',
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/admin/admin_attendance_course_detail.html', context)

@_admin_required
def admin_attendance_student_detail(request, student_id):
    from .models import User, AttendanceRecord
    student = get_object_or_404(User, id=student_id, role='student')
    
    courses_data = []
    for course in student.enrolled_courses.all():
        registers = course.attendance_registers.all()
        total_classes = registers.count()
        
        records = AttendanceRecord.objects.filter(student=student, register__course=course)
        presents = records.filter(status='present').count()
        absents = records.filter(status='absent').count()
        lates = records.filter(status='late').count()
        
        attended = presents + lates
        percentage = (attended / total_classes * 100) if total_classes > 0 else 0
        
        courses_data.append({
            'course': course,
            'total_classes': total_classes,
            'presents': presents,
            'absents': absents,
            'lates': lates,
            'percentage': round(percentage, 1),
        })
        
    records_history = AttendanceRecord.objects.filter(student=student).select_related('register', 'register__course').order_by('-register__date')
    
    context = {
        'student': student,
        'courses_data': courses_data,
        'records_history': records_history,
        'sidebar_active': 'asistencias',
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/admin/admin_attendance_student_detail.html', context)


# ============================================================
# ===  GESTIÓN DE COHORTES (Ciclo de vida de cursos)  ===
# ============================================================

@login_required
def admin_course_cohorts(request, course_id):
    """Lista todas las cohortes de un curso."""
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    course = get_object_or_404(Course, id=course_id)
    cohorts = course.cohorts.all()
    
    active_cohort = cohorts.filter(status='active').first()
    completed_cohorts = cohorts.filter(status='completed')
    
    context = {
        'course': course,
        'active_cohort': active_cohort,
        'completed_cohorts': completed_cohorts,
        'sidebar_active': 'cursos',
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/admin/admin_course_cohorts.html', context)


@login_required
def admin_cohort_create(request, course_id):
    """Crear una nueva cohorte para un curso."""
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    course = get_object_or_404(Course, id=course_id)
    
    # Verificar que no hay cohorte activa
    if course.cohorts.filter(status='active').exists():
        messages.error(request, 'Ya existe un grupo activo para este curso. Ciérralo antes de crear uno nuevo.')
        return redirect('admin_course_cohorts', course_id=course.id)
    
    if request.method == 'POST':
        form = CohortForm(request.POST, course=course)
        if form.is_valid():
            cohort = form.save(commit=False)
            cohort.course = course
            cohort.save()
            form.save_m2m()  # Guardar estudiantes
            
            # También agregar los estudiantes de la cohorte al curso si no están
            for student in cohort.students.all():
                course.students.add(student)
            
            messages.success(request, f'Grupo "{cohort.name}" creado exitosamente con {cohort.students.count()} estudiantes.')
            return redirect('admin_course_cohorts', course_id=course.id)
    else:
        from .models import PlatformSetting
        settings = PlatformSetting.get_settings()
        form = CohortForm(course=course, initial={'retention_months': settings.default_retention_months})
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Nuevo Grupo — {course.title}',
        'back_url': 'admin_course_cohorts',
        'back_id': course.id,
    })


@login_required
def admin_cohort_close(request, cohort_id):
    """Cerrar/finalizar una cohorte."""
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    cohort = get_object_or_404(Cohort, id=cohort_id, status='active')
    course = cohort.course
    
    if request.method == 'POST':
        form = CohortCloseForm(request.POST)
        if form.is_valid():
            # 1. Marcar cohorte como completada
            cohort.status = 'completed'
            cohort.completed_at = form.cleaned_data['completed_at']
            cohort.retention_months = int(form.cleaned_data['retention_months'])
            cohort.save()
            
            # 2. Vincular entregas existentes a esta cohorte
            cohort_students = cohort.students.all()
            Submission.objects.filter(
                assignment__module__course=course,
                student__in=cohort_students,
                cohort__isnull=True
            ).update(cohort=cohort)
            
            # 3. Remover estudiantes de la cohorte del curso activo
            for student in cohort_students:
                course.students.remove(student)
                
            # 4. Congelar el contenido del curso para esta cohorte
            from core.utils import freeze_cohort_content
            freeze_cohort_content(cohort)
            
            msg = f'Grupo "{cohort.name}" finalizado exitosamente.'
            messages.success(request, msg)
            return redirect('admin_course_cohorts', course_id=course.id)
    else:
        form = CohortCloseForm(initial={
            'completed_at': timezone.now().date(),
            'retention_months': cohort.retention_months,
        })
    
    return render(request, 'dashboards/teacher/teacher_form.html', {
        'form': form,
        'title': f'Cerrar Grupo: {cohort.name}',
        'back_url': 'admin_course_cohorts',
        'back_id': course.id,
        'submission': None,  # Para que use max-w-xl
    })


@login_required
def admin_cohort_detail(request, cohort_id):
    """Ver los detalles de una cohorte (archivada o activa)."""
    if request.user.role != 'admin': return redirect_dashboard_by_role(request.user)
    cohort = get_object_or_404(Cohort, id=cohort_id)
    course = cohort.course
    
    # Obtener las entregas de esta cohorte
    students_data = []
    for student in cohort.students.all().order_by('first_name', 'lastname'):
        submissions = Submission.objects.filter(
            student=student,
            assignment__module__course=course,
            cohort=cohort
        ).select_related('assignment')
        
        graded = submissions.filter(score__isnull=False)
        avg_score = graded.aggregate(avg=Avg('score'))['avg']
        
        students_data.append({
            'student': student,
            'submissions_count': submissions.count(),
            'graded_count': graded.count(),
            'avg_score': round(avg_score, 2) if avg_score else None,
        })
    
    context = {
        'cohort': cohort,
        'course': course,
        'students_data': students_data,
        'sidebar_active': 'cursos',
    }
    context.update(_sidebar_context(request))
    return render(request, 'dashboards/admin/admin_cohort_detail.html', context)
